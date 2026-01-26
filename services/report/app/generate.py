"""
AFASA 2.0 - Report Generator
PDF and Excel report generation
"""
from datetime import datetime, timezone
from typing import List, Dict, Any
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_pdf_report(
    tenant_name: str,
    date_from: datetime,
    date_to: datetime,
    summary: Dict[str, Any],
    detections: List[Dict[str, Any]],
    assessments: List[Dict[str, Any]],
    tasks: List[Dict[str, Any]]
) -> bytes:
    """Generate a PDF farm health report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    story.append(Paragraph("🌱 AFASA Farm Health Report", title_style))
    
    # Report info
    story.append(Paragraph(f"<b>Tenant:</b> {tenant_name}", styles['Normal']))
    story.append(Paragraph(
        f"<b>Period:</b> {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}",
        styles['Normal']
    ))
    story.append(Paragraph(
        f"<b>Generated:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
        styles['Normal']
    ))
    story.append(Spacer(1, 20))
    
    # Summary section
    story.append(Paragraph("📊 Summary", styles['Heading2']))
    summary_data = [
        ["Metric", "Value"],
        ["Total Snapshots", str(summary.get("total_snapshots", 0))],
        ["Total Detections", str(summary.get("total_detections", 0))],
        ["Assessments", str(summary.get("total_assessments", 0))],
        ["Open Tasks", str(summary.get("open_tasks", 0))],
        ["Completed Tasks", str(summary.get("completed_tasks", 0))]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Detections section
    if detections:
        story.append(Paragraph("🔍 Detection Summary", styles['Heading2']))
        det_data = [["Label", "Count", "Avg Confidence"]]
        
        label_stats = {}
        for det in detections:
            label = det.get("label", "unknown")
            if label not in label_stats:
                label_stats[label] = {"count": 0, "conf_sum": 0}
            label_stats[label]["count"] += 1
            label_stats[label]["conf_sum"] += det.get("confidence", 0)
        
        for label, stats in label_stats.items():
            avg_conf = stats["conf_sum"] / stats["count"] if stats["count"] > 0 else 0
            det_data.append([label, str(stats["count"]), f"{avg_conf:.1%}"])
        
        det_table = Table(det_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        det_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(det_table)
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    return buffer.getvalue()


def generate_xlsx_report(
    tenant_name: str,
    date_from: datetime,
    date_to: datetime,
    detections: List[Dict[str, Any]],
    assessments: List[Dict[str, Any]],
    tasks: List[Dict[str, Any]]
) -> bytes:
    """Generate an Excel report"""
    wb = Workbook()
    
    # Detections sheet
    ws_det = wb.active
    ws_det.title = "Detections"
    
    headers = ["ID", "Camera ID", "Label", "Confidence", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, det in enumerate(detections, 2):
        ws_det.cell(row=row, column=1, value=det.get("id", ""))
        ws_det.cell(row=row, column=2, value=det.get("camera_id", ""))
        ws_det.cell(row=row, column=3, value=det.get("label", ""))
        ws_det.cell(row=row, column=4, value=det.get("confidence", 0))
        ws_det.cell(row=row, column=5, value=det.get("created_at", ""))
    
    # Assessments sheet
    ws_ass = wb.create_sheet("Assessments")
    headers = ["ID", "Camera ID", "Severity", "Hypotheses", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws_ass.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1565C0", fill_type="solid")
    
    for row, ass in enumerate(assessments, 2):
        ws_ass.cell(row=row, column=1, value=ass.get("id", ""))
        ws_ass.cell(row=row, column=2, value=ass.get("camera_id", ""))
        ws_ass.cell(row=row, column=3, value=ass.get("severity", ""))
        ws_ass.cell(row=row, column=4, value=str(ass.get("hypotheses", [])))
        ws_ass.cell(row=row, column=5, value=ass.get("created_at", ""))
    
    # Tasks sheet
    ws_tasks = wb.create_sheet("Tasks")
    headers = ["ID", "Title", "Priority", "Status", "Due At"]
    for col, header in enumerate(headers, 1):
        cell = ws_tasks.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF8F00", fill_type="solid")
    
    for row, task in enumerate(tasks, 2):
        ws_tasks.cell(row=row, column=1, value=task.get("id", ""))
        ws_tasks.cell(row=row, column=2, value=task.get("title", ""))
        ws_tasks.cell(row=row, column=3, value=task.get("priority", 3))
        ws_tasks.cell(row=row, column=4, value=task.get("status", ""))
        ws_tasks.cell(row=row, column=5, value=task.get("due_at", ""))
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
